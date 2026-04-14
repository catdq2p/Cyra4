import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import datetime
from openpyxl import load_workbook

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TPCRA v3.0 Dashboard",
    page_icon="🔐",
    layout="wide",
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
.kpi-label { font-size: 11px; font-weight: 600; text-transform: uppercase;
             letter-spacing: 0.07em; color: #6c757d; margin-bottom: 4px; }
.kpi-value { font-size: 28px; font-weight: 600; line-height: 1.1; }
.tier-critical { color: #A32D2D; }
.tier-high     { color: #854F0B; }
.tier-medium   { color: #185FA5; }
.tier-low      { color: #3B6D11; }
.response-pill {
    display: inline-block; padding: 2px 10px; border-radius: 20px;
    font-size: 11px; font-weight: 600; letter-spacing: 0.03em;
}
.stTabs [data-baseweb="tab"] { font-size: 13px; padding: 8px 16px; }
div[data-testid="metric-container"] {
    background: #f8f9fa; border-radius: 10px; padding: 14px;
    border: 1px solid #e9ecef;
}
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
DOMAIN_MAP = {
    "A": "Organizational Management",
    "B": "Human Resource Management",
    "C": "Infrastructure Security",
    "D": "Data Protection",
    "E": "Access Management",
    "F": "Application Security",
    "G": "System Security",
    "H": "Email Security",
    "I": "Mobile Devices",
    "J": "Incident Response",
    "K": "Cloud Services",
    "L": "Business Continuity",
    "M": "Supply Chain & Physical Security",
    "N": "AI & Emerging Technology Risk",
}

RESP_COLORS = {
    "Yes":     "#639922",
    "No":      "#E24B4A",
    "Partial": "#EF9F27",
    "N/A":     "#B4B2A9",
}

RESP_PILL = {
    "Yes":     "background:#EAF3DE;color:#27500A",
    "No":      "background:#FCEBEB;color:#A32D2D",
    "Partial": "background:#FAEEDA;color:#633806",
    "N/A":     "background:#F1EFE8;color:#5F5E5A",
    "—":       "background:#F1EFE8;color:#888780",
}

TIER_COLORS = {
    "Critical": "#A32D2D",
    "High":     "#854F0B",
    "Medium":   "#185FA5",
    "Low":      "#3B6D11",
}

TIER_PILL = {
    "Critical": "background:#FCEBEB;color:#A32D2D",
    "High":     "background:#FAEEDA;color:#854F0B",
    "Medium":   "background:#E6F1FB;color:#0C447C",
    "Low":      "background:#EAF3DE;color:#3B6D11",
}

RATING_THRESHOLDS = [
    (90, "✅ Low",      "#639922"),
    (70, "🟡 Medium",   "#BA7517"),
    (50, "🟠 High",     "#D85A30"),
    (0,  "🔴 Critical", "#A32D2D"),
]

YES_VALS  = {"yes", "y"}
NO_VALS   = {"no", "n"}
NA_VALS   = {"n/a", "na", "not applicable"}
PART_VALS = {"partial", "partly", "partially"}
EVIDENCE_STATUS = {"submitted", "provided", "received", "complete", "done", "yes"}


# ── Helpers ────────────────────────────────────────────────────────────────────
def normalize_response(val) -> str:
    if val is None:
        return "—"
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%m/%d/%Y")
    s = str(val).strip().lower()
    if s in YES_VALS:  return "Yes"
    if s in NO_VALS:   return "No"
    if s in NA_VALS:   return "N/A"
    if s in PART_VALS: return "Partial"
    if not s or s == "—": return "—"
    return str(val).strip()


def extract_domain(key) -> str:
    if not key:
        return ""
    s = str(key).strip()
    # "A — ORGANIZATIONAL MANAGEMENT" style headers
    if " — " in s:
        return s.split(" — ")[0].strip().upper()
    if s and s[0].isalpha():
        return s[0].upper()
    return ""


def compliance_score(items: list) -> int:
    scored = [i for i in items if i["norm"] not in ("—", "N/A")]
    if not scored:
        return 0
    earned = sum(100 if i["norm"] == "Yes" else (50 if i["norm"] == "Partial" else 0)
                 for i in scored)
    return round(earned / len(scored))


def risk_rating(score: int) -> tuple:
    for threshold, label, color in RATING_THRESHOLDS:
        if score >= threshold:
            return label, color
    return "🔴 Critical", "#A32D2D"


def pill(text: str, style: str) -> str:
    return f'<span class="response-pill" style="{style}">{text}</span>'


def resp_pill(norm: str) -> str:
    return pill(norm if norm != "—" else "—", RESP_PILL.get(norm, RESP_PILL["—"]))


def tier_pill(tier: str) -> str:
    return pill(tier, TIER_PILL.get(tier, "background:#F1EFE8;color:#888780"))


# ── Parsers ────────────────────────────────────────────────────────────────────
def parse_part1(wb) -> dict:
    """Parse Part 1 — Contact & Engagement information."""
    if "Part 1" not in wb.sheetnames:
        return {}
    ws = wb["Part 1"]
    rows = list(ws.iter_rows(values_only=True))
    meta = {"title": "", "sections": {}, "items": []}

    current_section = ""
    for row in rows:
        if not any(v is not None for v in row):
            continue
        key, question, response = row[0], row[1], row[2] if len(row) > 2 else None

        # Title row
        if isinstance(key, str) and "TPCRA" in str(key) and not meta["title"]:
            meta["title"] = str(key).strip()
            continue

        # Section header
        if isinstance(key, str) and key.startswith("SECTION"):
            current_section = str(key).strip()
            meta["sections"].setdefault(current_section, [])
            continue

        # Data row
        if question and str(question).strip():
            item = {
                "key":      str(key).strip() if key else "",
                "section":  current_section,
                "question": str(question).strip(),
                "response": str(response).strip() if response else "",
                "other":    str(row[3]).strip() if len(row) > 3 and row[3] else "",
                "tier":     str(row[4]).strip() if len(row) > 4 and row[4] and str(row[4]) != "—" else "",
            }
            meta["items"].append(item)
            if current_section:
                meta["sections"][current_section].append(item)

    return meta


def parse_part2(wb) -> dict:
    """Parse Part 2 — Security questionnaire with responses, tiers, and remarks."""
    if "Part 2" not in wb.sheetnames:
        return {}
    ws = wb["Part 2"]
    rows = list(ws.iter_rows(values_only=True))

    result = {"title": "", "domains": {}, "items": []}
    current_domain = ""
    current_sub = ""

    for row in rows:
        if not any(v is not None for v in row):
            continue
        key = row[0]
        question  = row[1] if len(row) > 1 else None
        response  = row[2] if len(row) > 2 else None
        other     = row[3] if len(row) > 3 else None
        tier      = row[4] if len(row) > 4 else None

        key_s = str(key).strip() if key else ""

        # Title row
        if isinstance(key, str) and "TPCRA" in key and not result["title"]:
            result["title"] = key.strip()
            continue

        # Column header row
        if key_s == "#":
            continue

        # Domain header: "A — ORGANIZATIONAL MANAGEMENT"
        if isinstance(key, str) and " — " in key and key[0].isalpha() and len(key.split(" — ")[0]) == 1:
            letter = key.split(" — ")[0].strip().upper()
            name   = key.split(" — ", 1)[1].strip()
            current_domain = letter
            current_sub = ""
            result["domains"].setdefault(letter, {
                "name": DOMAIN_MAP.get(letter, name),
                "items": []
            })
            continue

        # Sub-section label rows (no response, key has a dot or is descriptive text)
        if isinstance(key, str) and question and response is None and tier is None:
            current_sub = str(question).strip() if question else ""
            continue

        # Question row — must have a proper key like A.1, B.2.3, N.1.1 etc.
        if key_s and question and str(question).strip():
            # Determine domain from key
            domain_letter = key_s[0].upper() if key_s[0].isalpha() else current_domain

            tier_s = str(tier).strip() if tier and str(tier).strip() not in ("—", "None", "") else ""
            norm   = normalize_response(response)

            item = {
                "key":      key_s,
                "domain":   domain_letter,
                "domain_name": DOMAIN_MAP.get(domain_letter, domain_letter),
                "sub":      current_sub,
                "question": str(question).strip(),
                "response": str(response).strip() if response else "",
                "norm":     norm,
                "other":    str(other).strip() if other else "",
                "tier":     tier_s,
            }

            result["items"].append(item)
            if domain_letter in result["domains"]:
                result["domains"][domain_letter]["items"].append(item)
            elif domain_letter:
                result["domains"].setdefault(domain_letter, {
                    "name": DOMAIN_MAP.get(domain_letter, domain_letter),
                    "items": [item]
                })

    return result


def parse_evidence(wb) -> list:
    """Parse Evidence checklist sheet."""
    if "Evidence" not in wb.sheetnames:
        return []
    ws = wb["Evidence"]
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for row in rows[2:]:
        if not any(v is not None for v in row):
            continue
        num, evidence, guidance, status, remarks, required_for = (
            row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else None
        )
        if evidence:
            status_norm = "Submitted" if str(status or "").strip().lower() in EVIDENCE_STATUS else (
                str(status).strip() if status else "Pending"
            )
            items.append({
                "num":          str(num).strip() if num else "",
                "evidence":     str(evidence).strip(),
                "guidance":     str(guidance).strip() if guidance else "",
                "status":       status_norm,
                "remarks":      str(remarks).strip() if remarks else "",
                "required_for": str(required_for).strip() if required_for else "",
            })
    return items


def extract_contact(p1_items: list) -> dict:
    """Pull key contact fields from Part 1 items."""
    contact = {"vendor": "", "rep": "", "email": "", "engagement": ""}
    for item in p1_items:
        q = item["question"].lower()
        r = item["response"]
        if not r or r in ("None", "—"):
            continue
        if "company name" in q:
            contact["vendor"] = r
        elif "authorized representative" in q and "email" not in q:
            contact["rep"] = r
        elif "email" in q and "representative" in q:
            contact["email"] = r
        elif "description of the engagement" in q:
            contact["engagement"] = r[:120] + "…" if len(r) > 120 else r
    return contact


def make_sample_excel() -> bytes:
    """Generate a minimal sample Excel matching v3.0 format."""
    p2_rows = [
        ("TPCRA Questionnaire - Part 2  |  v3.0  |  Response options: Yes / No / Partial / N/A", None, None, None, None, None),
        ("#", "Statement / Question", "Response\n(Yes/No/Partial/N/A)", "Other Information\n(Remarks & Evidence)", "Risk\nTier", "Comments\nRequired"),
        ("A — ORGANIZATIONAL MANAGEMENT", None, None, None, None, None),
        ("A.1", "IT Security policies and procedures are formally established and documented.", None, None, "Critical", "—"),
        ("A.2", "IT Security policies and procedures are reviewed at least annually.", None, None, "High", "—"),
        ("A.5", "IT Security policies are formally acknowledged by all employees and contractors.", None, None, "High", "—"),
        ("A.6", "IT Security policies comply with relevant regulatory requirements.", None, None, "Critical", "—"),
        ("B — HUMAN RESOURCE MANAGEMENT", None, None, None, None, None),
        ("B.2", "IT security awareness training is provided to ALL employees.", None, None, "High", "—"),
        ("B.3", "New hires are subjected to background screening and must sign an NDA.", None, None, "Critical", "—"),
        ("B.5", "A formal employee offboarding process revokes all access on the last working day.", None, None, "Critical", "—"),
        ("C — INFRASTRUCTURE SECURITY", None, None, None, None, None),
        ("C.1.2", "Clear network segmentation exists between web, application, and database tiers.", None, None, "Critical", "—"),
        ("C.1.5", "All inter-system communication uses encrypted protocols only (HTTPS, SFTP, TLS 1.2+).", None, None, "Critical", "—"),
        ("C.5.1", "All remote connections require MFA and encrypted communications.", None, None, "Critical", "—"),
        ("C.6.2", "Security patches are applied within defined SLA timelines based on criticality.", None, None, "Critical", "—"),
        ("C.7.2", "Anti-malware solutions are deployed on all user computers, servers, and endpoints.", None, None, "Critical", "—"),
        ("D — DATA PROTECTION", None, None, None, None, None),
        ("D.1.3", "Technical measures isolate company data from other clients.", None, None, "Critical", "—"),
        ("D.1.10", "A DLP solution or equivalent is deployed to prevent unauthorized data exfiltration.", None, None, "High", "—"),
        ("D.2.1", "Strong cryptographic protocols protect sensitive data in transit.", None, None, "Critical", "—"),
        ("E — ACCESS MANAGEMENT", None, None, None, None, None),
        ("E.1.1", "Access to critical systems is granted only on a need-to-know basis.", None, None, "Critical", "—"),
        ("E.1.7", "Sharing of user IDs or privileged accounts is strictly prohibited.", None, None, "Critical", "—"),
        ("E.1.8", "MFA is enforced for all remote access and privileged account usage.", None, None, "Critical", "—"),
        ("J — INCIDENT RESPONSE", None, None, None, None, None),
        ("J.2", "A documented Incident Response Plan (IRP) is established.", None, None, "Critical", "—"),
        ("J.3", "The IRP is tested at least annually.", None, None, "Critical", "—"),
        ("N — AI & EMERGING TECHNOLOGY RISK", None, None, None, None, None),
        ("N.1.1", "A formal AI usage policy is in place covering acceptable use and accountability.", None, None, "Critical", "—"),
        ("N.2.1", "Data will NOT be used to train AI models without explicit written consent.", None, None, "Critical", "—"),
    ]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df = pd.DataFrame(p2_rows)
        df.to_excel(writer, index=False, header=False, sheet_name="Part 2")
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.header("Upload questionnaire")
    uploaded = st.file_uploader(
        "TPCRA v3.0 Excel (.xlsx)",
        type=["xlsx", "xls"],
        help="Upload a completed TPCRA v3.0 questionnaire."
    )

# ── Empty state ────────────────────────────────────────────────────────────────
if not uploaded:
    st.title("🔐 TPCRA v3.0 Risk Assessment Dashboard")
    st.markdown("Upload a completed TPCRA questionnaire from the sidebar to generate the dashboard.")
    st.divider()
    c1, c2, c3, c4 = st.columns(4)
    c1.info("**Overview**\nCompliance score, risk rating, and response distribution across all 14 domains.")
    c2.info("**By domain**\nDrill into A–N domains with per-question response cards, tier badges, and remarks.")
    c3.info("**Gap analysis**\nAll No / Partial / unanswered items filtered by risk tier for prioritized remediation.")
    c4.info("**Evidence & Part 1**\nEvidence checklist status and engagement/contact information from Part 1.")
    st.stop()

# ── Load workbook ──────────────────────────────────────────────────────────────
try:
    wb = load_workbook(uploaded, read_only=True, data_only=True)
except Exception as e:
    st.error(f"Could not open file: {e}")
    st.stop()

p1_data  = parse_part1(wb)
p2_data  = parse_part2(wb)
evidence = parse_evidence(wb)

if not p2_data or not p2_data.get("items"):
    st.error("No Part 2 question data found. Ensure the file has a 'Part 2' sheet matching the TPCRA v3.0 format.")
    st.stop()

contact  = extract_contact(p1_data.get("items", []))
p2_items = p2_data["items"]
domains  = p2_data["domains"]

# ── Header ─────────────────────────────────────────────────────────────────────
h1, h2 = st.columns([4, 1])
with h1:
    vendor_label = contact["vendor"] or "Vendor"
    st.title(f"🔐 {vendor_label}")
    if contact["rep"] or contact["email"]:
        st.caption(f"{contact['rep']}  ·  {contact['email']}")
    if contact["engagement"]:
        st.caption(f"Engagement: {contact['engagement']}")
with h2:
    st.caption(p2_data.get("title", "TPCRA v3.0"))

st.divider()

# ── KPI metrics ────────────────────────────────────────────────────────────────
answered = [i for i in p2_items if i["norm"] in ("Yes", "No", "Partial", "N/A")]
unanswered = [i for i in p2_items if i["norm"] == "—"]
n_yes  = sum(1 for i in p2_items if i["norm"] == "Yes")
n_no   = sum(1 for i in p2_items if i["norm"] == "No")
n_part = sum(1 for i in p2_items if i["norm"] == "Partial")
n_na   = sum(1 for i in p2_items if i["norm"] == "N/A")
n_unans = len(unanswered)
score  = compliance_score(p2_items)
rating_label, rating_color = risk_rating(score)

k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
k1.metric("Total questions", len(p2_items))
k2.metric("✅ Yes",           n_yes)
k3.metric("❌ No",            n_no)
k4.metric("⚠️ Partial",       n_part)
k5.metric("➖ N/A",           n_na)
k6.metric("⬜ Unanswered",    n_unans)
k7.metric("Compliance score",  f"{score}%",
    delta=rating_label,
    delta_color="normal" if score >= 70 else ("off" if score >= 50 else "inverse"),
)
st.divider()

# ── Tabs ───────────────────────────────────────────────────────────────────────
tab_overview, tab_domain, tab_gaps, tab_evidence, tab_part1 = st.tabs([
    "Overview", "By domain", "Gap analysis", "Evidence checklist", "Engagement Info"
])

# ══════════════════════════
# TAB 1 — OVERVIEW
# ══════════════════════════
with tab_overview:
    col_bar, col_right = st.columns([3, 2])

    with col_bar:
        st.subheader("Compliance by domain")
        dom_rows = []
        for letter, dom in domains.items():
            items = dom["items"]
            if not items:
                continue
            sc = compliance_score(items)
            rl, rc = risk_rating(sc)
            dom_rows.append({
                "Domain":  dom["name"],
                "Yes":     sum(1 for i in items if i["norm"] == "Yes"),
                "Partial": sum(1 for i in items if i["norm"] == "Partial"),
                "No":      sum(1 for i in items if i["norm"] == "No"),
                "N/A":     sum(1 for i in items if i["norm"] == "N/A"),
                "Score":   sc,
                "Rating":  rl,
            })
        dom_df = pd.DataFrame(dom_rows)

        if not dom_df.empty:
            melt = dom_df.melt(
                id_vars="Domain", value_vars=["Yes", "Partial", "No", "N/A"],
                var_name="Response", value_name="Count"
            ).query("Count > 0")
            fig_bar = px.bar(
                melt, x="Count", y="Domain", color="Response",
                orientation="h",
                color_discrete_map=RESP_COLORS,
                category_orders={"Response": ["Yes", "Partial", "No", "N/A"]},
                labels={"Count": "Questions", "Domain": ""},
                height=max(380, len(dom_df) * 42),
            )
            fig_bar.update_layout(
                margin=dict(l=0, r=0, t=10, b=0),
                legend_title_text="Response",
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                yaxis={"categoryorder": "total ascending"},
                font=dict(size=12),
            )
            fig_bar.update_xaxes(showgrid=True, gridcolor="rgba(0,0,0,0.07)")
            st.plotly_chart(fig_bar, use_container_width=True)

    with col_right:
        st.subheader("Distribution")
        labels = [k for k, v in [("Yes",n_yes),("Partial",n_part),("No",n_no),("N/A",n_na)] if v > 0]
        values = [v for k, v in [("Yes",n_yes),("Partial",n_part),("No",n_no),("N/A",n_na)] if v > 0]
        fig_donut = go.Figure(go.Pie(
            labels=labels, values=values, hole=0.62,
            marker_colors=[RESP_COLORS[l] for l in labels],
            textinfo="label+percent",
            hovertemplate="%{label}: %{value}<extra></extra>",
        ))
        fig_donut.update_layout(
            margin=dict(l=0, r=0, t=10, b=0), showlegend=False,
            height=260, paper_bgcolor="rgba(0,0,0,0)", font=dict(size=12),
        )
        st.plotly_chart(fig_donut, use_container_width=True)

        st.subheader("Domain scores")
        if not dom_df.empty:
            for _, row in dom_df.sort_values("Score").iterrows():
                _, rc = risk_rating(row["Score"])
                st.markdown(
                    f'<div style="margin-bottom:7px">'
                    f'<div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px">'
                    f'<span style="color:#555">{row["Domain"]}</span>'
                    f'<span style="font-weight:600;color:{rc}">{row["Score"]}%</span></div>'
                    f'<div style="background:#f0f0f0;border-radius:4px;height:5px;overflow:hidden">'
                    f'<div style="width:{row["Score"]}%;height:100%;background:{rc};border-radius:4px"></div>'
                    f'</div></div>', unsafe_allow_html=True,
                )

    # Risk tier breakdown
    st.divider()
    st.subheader("Gaps by risk tier")
    gaps = [i for i in p2_items if i["norm"] in ("No", "Partial", "—") and i["tier"]]
    tier_counts = {}
    for g in gaps:
        tier_counts[g["tier"]] = tier_counts.get(g["tier"], 0) + 1

    if tier_counts:
        t_labels = ["Critical", "High", "Medium", "Low"]
        t_vals   = [tier_counts.get(t, 0) for t in t_labels]
        t_colors = [TIER_COLORS.get(t, "#888") for t in t_labels]
        fig_tier = go.Figure(go.Bar(
            x=t_labels, y=t_vals,
            marker_color=t_colors,
            text=t_vals, textposition="outside",
        ))
        fig_tier.update_layout(
            margin=dict(l=0, r=0, t=10, b=0), height=200,
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            showlegend=False, font=dict(size=12),
            yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.07)"),
        )
        st.plotly_chart(fig_tier, use_container_width=True)
    else:
        st.info("No gap data with risk tier information found.")


# ══════════════════════════
# TAB 2 — BY DOMAIN
# ══════════════════════════
with tab_domain:
    domain_choices = [f"{l} — {domains[l]['name']}" for l in domains if domains[l]["items"]]
    if not domain_choices:
        st.info("No domain data found.")
    else:
        chosen = st.selectbox("Select domain", domain_choices)
        chosen_letter = chosen.split(" — ")[0]
        dom = domains[chosen_letter]
        items = dom["items"]

        sc = compliance_score(items)
        rl, rc = risk_rating(sc)
        n_y = sum(1 for i in items if i["norm"] == "Yes")
        n_n = sum(1 for i in items if i["norm"] == "No")
        n_p = sum(1 for i in items if i["norm"] == "Partial")
        n_a = sum(1 for i in items if i["norm"] == "N/A")

        dm1, dm2, dm3, dm4, dm5, dm6 = st.columns(6)
        dm1.metric("Questions", len(items))
        dm2.metric("✅ Yes",    n_y)
        dm3.metric("❌ No",     n_n)
        dm4.metric("⚠️ Partial", n_p)
        dm5.metric("➖ N/A",    n_a)
        dm6.metric("Score", f"{sc}%", delta=rl,
            delta_color="normal" if sc >= 70 else ("off" if sc >= 50 else "inverse"))

        st.divider()

        # Filter
        resp_f = st.multiselect(
            "Filter by response", ["Yes","No","Partial","N/A","—"],
            default=["Yes","No","Partial","N/A","—"], key="dom_resp_filter"
        )
        shown = [i for i in items if i["norm"] in resp_f]
        st.caption(f"Showing {len(shown)} of {len(items)} questions")

        current_sub = ""
        for item in shown:
            # Sub-section label
            if item["sub"] and item["sub"] != current_sub:
                current_sub = item["sub"]
                st.markdown(
                    f'<div style="margin:16px 0 6px;font-size:11px;font-weight:700;'
                    f'text-transform:uppercase;letter-spacing:0.06em;color:#6c757d">'
                    f'{current_sub}</div>', unsafe_allow_html=True
                )

            norm = item["norm"]
            bg = {"No": "#fff5f5", "Partial": "#fffaf5"}.get(norm, "transparent")
            border = {"No": "#f7c1c1", "Partial": "#FAC775"}.get(norm, "#e9ecef")
            tier_badge = tier_pill(item["tier"]) if item["tier"] else ""
            resp_badge = resp_pill(norm)
            key_line = item["key"]
            tier_part = "&nbsp;&nbsp;" + tier_badge if tier_badge else ""
            remarks_html = (
                '<div style="font-size:12px;color:#555;margin-top:6px;'
                'padding-top:6px;border-top:0.5px solid #e9ecef">'
                + item["other"] + "</div>"
            ) if item["other"] else ""
            card_html = (
                '<div style="padding:10px 14px;border-radius:8px;margin-bottom:6px;'
                'background:' + bg + ';border:0.5px solid ' + border + '">'
                '<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:10px">'
                '<div style="flex:1">'
                '<div style="font-size:11px;color:#aaa;margin-bottom:3px">'
                + key_line + tier_part +
                '</div>'
                '<div style="font-size:13px;color:#333;line-height:1.55">' + item["question"] + '</div>'
                + remarks_html +
                '</div>'
                '<div style="flex-shrink:0;padding-top:2px">' + resp_badge + '</div>'
                '</div>'
                '</div>'
            )
            st.markdown(card_html, unsafe_allow_html=True)

            # Long free-text responses
            if item["response"] and len(item["response"]) > 50 and norm not in ("Yes","No","Partial","N/A","—"):
                with st.expander(f"View full response — {item['key']}"):
                    st.markdown(
                        f'<div style="font-size:13px;color:#444;line-height:1.6;white-space:pre-wrap">'
                        f'{item["response"]}</div>', unsafe_allow_html=True
                    )


# ══════════════════════════
# TAB 3 — GAP ANALYSIS
# ══════════════════════════
with tab_gaps:
    gaps_all = [i for i in p2_items if i["norm"] in ("No", "Partial", "—")]
    n_crit = sum(1 for g in gaps_all if g["tier"] == "Critical")
    n_high = sum(1 for g in gaps_all if g["tier"] == "High")
    n_med  = sum(1 for g in gaps_all if g["tier"] == "Medium")
    n_low  = sum(1 for g in gaps_all if g["tier"] == "Low")

    if not gaps_all:
        st.success("No gaps found — all questions are compliant or marked N/A.")
    else:
        g1, g2, g3, g4, g5 = st.columns(5)
        g1.metric("Total gaps",        len(gaps_all))
        g2.metric("🔴 Critical tier",  n_crit)
        g3.metric("🟠 High tier",      n_high)
        g4.metric("🟡 Medium tier",    n_med)
        g5.metric("🟢 Low tier",       n_low)
        st.divider()

        # Filters
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            tier_f = st.multiselect(
                "Filter by risk tier",
                ["Critical","High","Medium","Low",""],
                default=["Critical","High","Medium","Low",""],
                format_func=lambda x: x if x else "No tier",
            )
        with fc2:
            resp_f2 = st.multiselect(
                "Filter by response",
                ["No","Partial","—"],
                default=["No","Partial","—"],
                format_func=lambda x: x if x != "—" else "Unanswered",
            )
        with fc3:
            dom_f = st.multiselect(
                "Filter by domain",
                sorted({g["domain_name"] for g in gaps_all}),
                default=sorted({g["domain_name"] for g in gaps_all}),
            )

        shown_gaps = [
            g for g in gaps_all
            if g["tier"] in tier_f and g["norm"] in resp_f2 and g["domain_name"] in dom_f
        ]

        # Sort: Critical first
        tier_order = {"Critical":0,"High":1,"Medium":2,"Low":3,"":4}
        shown_gaps.sort(key=lambda x: tier_order.get(x["tier"], 4))

        st.caption(f"Showing {len(shown_gaps)} of {len(gaps_all)} gaps")

        for g in shown_gaps:
            norm  = g["norm"]
            bg    = {"No": "#fff5f5", "Partial": "#fffaf5", "—": "#fafafa"}.get(norm, "#fafafa")
            border = {"No": "#f7c1c1", "Partial": "#FAC775", "—": "#e9ecef"}.get(norm, "#e9ecef")
            t_badge = tier_pill(g["tier"]) if g["tier"] else ""
            r_badge = resp_pill(norm)
            unans_note = '<span style="font-size:11px;color:#aaa"> — not yet answered</span>' if norm == "—" else ""

            st.markdown(
                f'<div style="padding:10px 14px;border-radius:8px;margin-bottom:6px;'
                f'background:{bg};border:0.5px solid {border}">'
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:10px">'
                f'<div style="flex:1">'
                f'<div style="font-size:11px;color:#aaa;margin-bottom:3px">'
                f'{g["key"]} &nbsp;·&nbsp; {g["domain_name"]}'
                f'{("&nbsp;&nbsp;" + t_badge) if t_badge else ""}'
                f'</div>'
                f'<div style="font-size:13px;color:#333;line-height:1.55">'
                f'{g["question"]}{unans_note}</div>'
                f'</div>'
                f'<div style="flex-shrink:0;padding-top:2px">{r_badge}</div>'
                f'</div>'
                f'</div>', unsafe_allow_html=True,
            )

        # Export
        st.divider()
        gap_df = pd.DataFrame([{
            "Key":       g["key"],
            "Domain":    g["domain_name"],
            "Question":  g["question"],
            "Response":  g["norm"],
            "Risk Tier": g["tier"],
            "Remarks":   g["other"],
        } for g in shown_gaps])

        e1, e2, _ = st.columns([1,1,3])
        with e1:
            st.download_button(
                "⬇ Export gaps CSV", data=gap_df.to_csv(index=False).encode(),
                file_name="tpcra_gaps.csv", mime="text/csv", use_container_width=True,
            )
        with e2:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                gap_df.to_excel(writer, index=False, sheet_name="Gaps")
            st.download_button(
                "⬇ Export gaps Excel", data=buf.getvalue(),
                file_name="tpcra_gaps.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# ══════════════════════════
# TAB 4 — EVIDENCE
# ══════════════════════════
with tab_evidence:
    if not evidence:
        st.info("No Evidence sheet found in the uploaded file.")
    else:
        submitted = sum(1 for e in evidence if e["status"] == "Submitted")
        pending   = len(evidence) - submitted

        ev1, ev2, ev3 = st.columns(3)
        ev1.metric("Total evidence items", len(evidence))
        ev2.metric("✅ Submitted",          submitted)
        ev3.metric("⏳ Pending",            pending)
        st.divider()

        rows_html = ""
        for ev in evidence:
            s = ev["status"]
            s_style = "background:#EAF3DE;color:#27500A" if s == "Submitted" else "background:#FAEEDA;color:#633806"
            rows_html += (
                "<tr>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#aaa;width:4%;vertical-align:top">' + str(ev['num']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#333;width:26%;vertical-align:top;font-weight:500">' + str(ev['evidence']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#555;width:35%;vertical-align:top;line-height:1.5">' + str(ev['guidance']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;width:20%;vertical-align:top">' + str(ev['required_for']) + "</td>"
                '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;width:15%;vertical-align:top">'
                '<span class="response-pill" style="' + s_style + '">' + s + "</span></td>"
                "</tr>"
            )

        TH = "padding:9px 10px;text-align:left;font-size:11px;font-weight:600;color:#6c757d;text-transform:uppercase;letter-spacing:0.05em;border-bottom:1px solid #e9ecef"
        ev_table = (
            '<div style="border:1px solid #e9ecef;border-radius:10px;overflow:hidden">'
            '<table style="width:100%;border-collapse:collapse;table-layout:fixed">'
            '<thead><tr style="background:#f8f9fa">'
            '<th style="' + TH + ';width:4%">#</th>'
            '<th style="' + TH + ';width:26%">Evidence required</th>'
            '<th style="' + TH + ';width:35%">Guidance</th>'
            '<th style="' + TH + ';width:20%">Required for</th>'
            '<th style="' + TH + ';width:15%">Status</th>'
            "</tr></thead>"
            "<tbody>" + rows_html + "</tbody>"
            "</table></div>"
        )
        st.markdown(ev_table, unsafe_allow_html=True)


# ══════════════════════════
# TAB 5 — ENGAGEMENT INFO
# ══════════════════════════
with tab_part1:
    if not p1_data or not p1_data.get("items"):
        st.info("No Part 1 data found in the uploaded file. Ensure the file contains a 'Part 1' sheet.")
    else:
        p1_sections = p1_data.get("sections", {})
        if not p1_sections:
            sections_to_show = {"All questions": p1_data["items"]}
        else:
            sections_to_show = p1_sections

        for sec_name, sec_items in sections_to_show.items():
            if not sec_items:
                continue
            st.subheader(sec_name.replace("SECTION ", "").replace(" — ", " — ").title()
                         if "SECTION" in sec_name else sec_name)
            rows_html = ""
            for item in sec_items:
                r = item["response"] or "—"
                resp_preview = r if len(r) <= 100 else r[:97] + "…"
                tier_badge = tier_pill(item["tier"]) if item.get("tier") else ""
                rows_html += (
                    "<tr>"
                    '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#aaa;width:7%;vertical-align:top">' + item['key'] + "</td>"
                    '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#333;width:48%;vertical-align:top;line-height:1.5">' + item['question'] + "</td>"
                    '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;font-size:13px;color:#555;width:35%;vertical-align:top;line-height:1.5">' + resp_preview + "</td>"
                    '<td style="padding:9px 10px;border-bottom:1px solid #f0f0f0;width:10%;vertical-align:top">' + tier_badge + "</td>"
                    "</tr>"
                )

            TH2 = "padding:9px 10px;text-align:left;font-size:11px;font-weight:600;color:#6c757d;text-transform:uppercase;letter-spacing:0.05em;border-bottom:1px solid #e9ecef"
            p1_table = (
                '<div style="border:1px solid #e9ecef;border-radius:10px;overflow:hidden;margin-bottom:1.5rem">'
                '<table style="width:100%;border-collapse:collapse;table-layout:fixed">'
                '<thead><tr style="background:#f8f9fa">'
                '<th style="' + TH2 + ';width:7%">#</th>'
                '<th style="' + TH2 + ';width:48%">Question</th>'
                '<th style="' + TH2 + ';width:35%">Response</th>'
                '<th style="' + TH2 + ';width:10%">Tier</th>'
                "</tr></thead>"
                "<tbody>" + rows_html + "</tbody>"
                "</table></div>"
            )
            st.markdown(p1_table, unsafe_allow_html=True)

st.divider()
st.caption("TPCRA v3.0 — Third-Party Cyber Risk Assessment Dashboard  ·  For internal use only")
